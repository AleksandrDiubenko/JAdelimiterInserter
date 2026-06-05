from __future__ import annotations

import io
import subprocess
import sys
import unicodedata
from pathlib import Path
from typing import Any


try:
    import regex as re
except ModuleNotFoundError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "regex", "--quiet"])
    import regex as re

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    from openpyxl import load_workbook


DEFAULT_DELIMITER = "\u200B"
WORD_JOINER = "\u2060"
WORD_JOINER_BREAK_MARKER = "\uE000JDI_BREAK\uE000"
TARGET_HEADERS = {"ja", "jp", "jap", "japanese", "日本語", "日语"}
WHITESPACE_CHARS = {" ", "\t", "\n", "\r", "\v", "\f", "\u3000"}
JAPANESE_JOINABLE_PUNCTUATION = set("、。！？；：…「」『』［］")
NEXT_CHAR_BLOCKERS_RE = re.compile(r'[、。？！,．,.!?"”」』）)]')
PUNCTUATION_ONLY_RE = re.compile(r"^[、。？！…‥！？\s]*$")
LEADING_PUNCTUATION = "、。？！：；…‥" + "..."


def get_colab_files() -> Any | None:
    try:
        from google.colab import files
    except ModuleNotFoundError:
        return None
    return files


RuleGroup = tuple[str, tuple[str, ...]]

KATAKANA_CHAR_RE = r"\p{scx=Katakana}"
KATAKANA_CAPTURE_RE = rf"(?P<Kata>{KATAKANA_CHAR_RE}+)"
HAN_OR_FRESH_KATAKANA_RE = rf"(\p{{Han}}|{KATAKANA_CHAR_RE})"
CONTENT_START_RE = rf"(\p{{Han}}|{KATAKANA_CHAR_RE}|[0-9０-９])"
HAN_OR_KATA2_RE = rf"(\p{{Han}}{{2}}|{KATAKANA_CHAR_RE})"
JAPANESE_JOINABLE_CHAR_RE = re.compile(rf"(?:\p{{Han}}|\p{{Hiragana}}|{KATAKANA_CHAR_RE})$")


def before_content(rule: str) -> str:
    return rf"{rule}(?={CONTENT_START_RE})"


def before_han_or_katakana(rule: str) -> str:
    return rf"{rule}(?={HAN_OR_KATA2_RE})"


def any_of(rules: tuple[str, ...]) -> str:
    return "|".join(rules)


PARTICLE_BOUND_STEMS = (
    r"\p{Han}{1,2}",
    KATAKANA_CAPTURE_RE,
    r"こと",
    r"ところ",
    r"つもり",
    r"\p{Han}(?:\p{Hiragana}(?!で))+\p{Han}",
    r"(?P=Kata)\p{Han}",
    r"もの",
    r"入り",
    r"」",
    r"たち",
    r"ここ",
    r"そこ",
    r"\p{Han}ら",
    r"(?P<double>\p{Hiragana}{2})(?P=double)",
    r"[えけげせぜてでねめれ]る(?!べき)",
    r"まま",
    r"[あこそ]いつ",
    r"あ[なん]た",
    r"さん",
    r"まみれ",
    r"おそらく",
    r"たっぷり",
    r"気持ち",
    r"すら",
    r"さすが",
    r"くず",
    r"あちこち",
    r"もと",
    r"さま",
    r"[こそあど]れ",
    r"ど[れん]だけ",
    r"みんな",
    r"やつ",
    r"すで",
    r"とき",
    r"だ",
    r"[こそあ]ちら",
    r"[こそあ]っち",
    r"あるの"
    r"[あわ]たく?し",
    r"みたい",
    r"どこ",
    r"[0-9０-９][%％年月日個本人枚匹頭羽冊台隻つ]?",
    r"かり",
    r"\p{Han}[いきぎしじちにひみり](?!がけ)",
    r"のみ",
)

PARTICLE_BOUND_ENDINGS = (
    r"が(?!(して|った|[らりるれろ]|かり))",
    r"か(?!([はもらなえがけげせぜてでねめれいきぎしちにんをうくぐすつぬむるりっ]|った|さ))",
    r"か[は]",
    r"は(?!ず)",
    r"も(?![のろ])",
    r"の(?![みにがはた為よ])",
    r"なく(?!て)",
    r"な(?![くのんらるいし])",
    r"する(?![なのよ])",
    r"から(?!して)",
    r"まで",
    r"にとっては?",
    r"に(?!([はもね]|ついて|よって|よる|関して|すら|とって))",
    r"に[はも]",
    r"へ[のと]",
    r"へ(?![のと])",
    r"で(?![はもすしきの])",
    r"で[はも]",
    r"じゃあ",
    r"じて(?!る)",
    r"や(?![からりるれ])",
    r"と[のはか]",
    r"と(?!(し始め|して|[のなはかす]|[い言云]う))",
    r"して[はも]",
    r"して(?![はもる])",
    r"ならば",
    r"なら(?![ばで])",
)

PARTICLE_BOUND_RULE = rf"({any_of(PARTICLE_BOUND_STEMS)})({any_of(PARTICLE_BOUND_ENDINGS)})"

PUNCTUATION_RULES = (
    r"[、。？！・；]",
    r"(――)",
    r"(……)",
    r"(\.\.\.)",
)

TOPIC_AND_CONNECTOR_RULES = (
    r"について[はも]?",
    r"に関して[はも]?",
    rf"から(?={HAN_OR_FRESH_KATAKANA_RE})",
    r"[っいきぎしちにん][ただ]り",
    r"とにかく",
    r"でも",
    r"[くぐ]らいは?",
    r"まるで",
    r"(?<!と)って(?![るたかも])",
    r"っても",
    r"すなわち",
    r"つまり",
)

PARTICLE_AND_TAIL_RULES = (
    r"[うくぐすつぬふむる]の[にはもがをでよ](?!しょう)",
    r"を",
    r"んな[のに]",
    r"ったの[はが]",
    r"[って]たら",
    r"じゃ(?=なければ|なけりゃ)",
    r"として(?!も)",
    r"[ただ]と(?!(は|えば))",
    r"[ただ]とは",
    r"とは",
    r"だけで[はも]",
    r"だけで(?![はも])",
)

ADVERBIAL_AND_REPETITION_RULES = (
    r"ちょっと",
    r"ちょうど",
    r"々な",
    r"々に(?![もは])",
    r"々に[もは]",
    r"たい(?=\p{Han})",
    r"けど",
    r"よう[なに]?(?=(\p{Han}{2}|\p{scx=Katakana}))",
)

COMPARISON_AND_EVALUATION_RULES = (
    r"[のただ]ほうが",
    r"ないほうが",
    r"[のただ]方が",
    r"ない方が",
    r"風に",
    r"ほとんど",
    r"らしくて(?!は)",
    r"らしく(?!て)",
    r"ほうが(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"より(?=ずっと)",
    r"よりも",
    r"もっとも",
    r"かなり(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"よりかは",
    r"とっ?ても",
)

CONTINUATION_AND_RESULT_RULES = (
    r"[いきしちにひみり]たくて",
    r"[うくすつぬふむる]まて",
    r"[^一-龯]続く",
    r"く(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"すれば",
    r"て(?=い?ました)",
    r"しっかり",
    r"して(?=あげ([るた]|(ます|まし)))",
    r"て(?=(ください|下さい|ちょうだい))",
    r"[てで](?=くれ)",
    r"くなって(?!は)",
    r"され[るた](?![んの])",
    r"かった(?![んのりわっがぞよぜ])",
    r"もなくて(?!は)",
    r"もなく(?!て)",
    r"った(?![らんのりわっがぞよぜか])",
    r"した(?=(\p{Han}{2}|こと|とこ))",
    r"れて(?=(いき?ま|いる|いた|いな))",
    r"[えけげせぜてでねめれ]なく(?!て)",
    r"[えけげせぜてでねめれあかさたなまら]ずに",
    r"て(?=いな)",
    r"[えけげせぜてでねめれいきしじちにみりっ]て(?=い(る|ま|く|け))",
    r"\p{Han}し?い(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))(?!出)",
    r"\p{Han}し?くは(?=ない)",
    r"\p{Han}しく(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"[えけげせてねめれ]ば(?=(\p{Han}|\p{scx=Katakana}))",
    r"[えけげせぜてでねめれ]ば(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"りゃ(?=(\p{Han}|\p{scx=Katakana}))",
    r"[えけげせぜてでねめれ][るてた](?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"たく(?=な[いか])",
    r"[わかさたなまら]れ[るた](?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"\p{Han}ても",
    r"して(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"\p{Han}たる(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"な[くい](?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"[いきっ]た(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"て(?=ほし[いくか])",
    r"くて(?=\p{Han})",
    r"て(?=しま[ういわ])",
    r"とって(?!は)",
    r"[てで](?=しま)",
    r"[てで](?=お[かきくけこ])",
    r"[てで](?=みた[いか])",
)

CONDITIONAL_AND_NEGATIVE_RULES = (
    r"ないと(?=いけ)",
    r"(れば|ないと)(?=([い良善好]い|[よ良善好]か))",
    r"かも(?=[しれ])",
    r"どうなるか(?!は)",
    r"どうなるかは",
    r"[らば]いいの?か[はが]?",
    r"べきじゃ(?!あ)",
    r"(なければ|なきゃ|ないと)(?=(なら|いけ))",
    r"\p{Hiragana}(?=はず(だ|よ|$|。|…|！|？))",
    r"なら(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"ないと(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"しか(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
)

TEMPORAL_AND_SCOPE_RULES = (
    r"ため([にの](?![はも])|ならば|なら(?!ば))",
    r"ため[にの][はも]",
    r"為に(?![はも])",
    r"為に[はも]",
    r"いきなり",
    r"これまでに(?!は)",
    r"はじめて",
    r"あらゆる",
    r"すべて(の|を|では|じゃ(?!あ))",
    r"すぐに[はも]",
    r"すぐに(?![はも])",
    r"ながら",
    r"がてら",
    r"とともに(?![はも])",
    r"と共に(?![はも])",
    r"すべて(?![でのを])",
    r"ただの",
    r"まま(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"しばらく",
    r"ゆっくり(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"ちゃんと(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"いくつか",
)

PHRASE_AND_CLAUSE_RULES = (
    r"わけ(には|では|じゃ(?!あ))",
    r"どうしても?",
    r"どうやって",
    r"どうに[かも](?=(\p{Han}|\p{scx=Katakana}))",
    r"のもとに",
    r"[うくすつぬふむるじのい]よう[にな]",
    r"じゃ(?=な[いか])",
    r"では(?=な[いか])",
    r"またしても",
    r"なのは",
    r"という(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"\p{Han}\p{Hiragana}に(?=な(る|った|らな))",
    r"なんて(?=こった)",
    rf"による(?={HAN_OR_FRESH_KATAKANA_RE})",
    r"何かが",
    r"[うくぐすずつぬむる]には",
    r"そうで(?=\p{Han})",
)

SCRIPT_BOUNDARY_RULES = (
    r"\p{Hiragana}(?=(?P=Kata))",
    r"て(?=ありがと)",
    r"\p{Han}{2}(?=(?P=Kata))",
    r"な(?=(\p{Han}|\p{scx=Katakana}|[0-9０-９]))",
    r"(?P=Kata)(?=\p{Han}{2})",
    r"(?P<doubler>\p{Hiragana}{2})(?P=doubler)",
    r"\p{Han}\p{Hiragana}(?=\p{Han}{2})",
    r"とか(?=\p{Han})",
    r"もう(?=\p{Han})",
    r"\p{Hiragana}(?=つもり)",
    r"が(?=(\p{Han}{2}|\p{scx=Katakana}))",
)


# Add edge cases by placing a small regex in the narrowest matching group below.
# Keep groups ordered from specific/high-value rules to broader fallback rules:
# earlier alternatives can win when two matches start at the same position.
BREAK_RULE_GROUPS: tuple[RuleGroup, ...] = (
    ("copula negative guards", (r"\p{Han}じゃ(?=なければ|なけりゃ)",)),
    ("fixed compound particles", (r"にとっては?",)),
    ("conditional question compounds", (r"(すれば|れば|ったら|たら|[らば])いいの?か[はが]?",)),
    ("fixed question phrases", (r"誰かが", r"ど[れん]だけ[のかはも]?", r"するのか[がはもの]?", r"[えけげせてでねめれ]るの\u200B?も")),
    (
        "particle-bound phrases",
        (PARTICLE_BOUND_RULE,),
    ),
    ("punctuation", PUNCTUATION_RULES),
    ("degree suffixes", (r"\p{Han}すぎ[^たるだて]",)),
    ("topics and connectors", TOPIC_AND_CONNECTOR_RULES),
    ("particles and clause tails", PARTICLE_AND_TAIL_RULES),
    ("adverbials and repetition", ADVERBIAL_AND_REPETITION_RULES),
    ("comparison and evaluation", COMPARISON_AND_EVALUATION_RULES),
    ("continuation and result forms", CONTINUATION_AND_RESULT_RULES),
    ("conditionals and negatives", CONDITIONAL_AND_NEGATIVE_RULES),
    ("temporal and scope markers", TEMPORAL_AND_SCOPE_RULES),
    ("phrases and clause forms", PHRASE_AND_CLAUSE_RULES),
    ("script boundary fallbacks", SCRIPT_BOUNDARY_RULES),
)


def iter_break_rules(rule_groups: tuple[RuleGroup, ...] = BREAK_RULE_GROUPS):
    for _group_name, rules in rule_groups:
        yield from rules


def expand_legacy_rule_aliases(rule: str) -> str:
    return rule


def compile_break_pattern(rule_groups: tuple[RuleGroup, ...] = BREAK_RULE_GROUPS):
    rules = "\n    |\n    ".join(expand_legacy_rule_aliases(rule) for rule in iter_break_rules(rule_groups))
    return re.compile(f"(\n    {rules}\n)", re.VERBOSE)


BREAK_PATTERN = compile_break_pattern()

def postprocess_ellipses(text: str, delimiter: str) -> str:
    text = re.sub(rf"^(…{{1,4}}){re.escape(delimiter)}", r"\1", text)
    text = re.sub(r"(?<!…)(…)(?!…)(?=\S)", lambda match: match.group(1) + delimiter, text)
    text = re.sub(rf"([^\s…]){re.escape(delimiter)}(…|\.\.\.)", r"\1\2", text)
    return text


def should_insert_after_match(text: str, match_end: int) -> bool:
    remainder = text[match_end:]
    if not remainder:
        return False
    if NEXT_CHAR_BLOCKERS_RE.match(remainder[0]):
        return False
    return not PUNCTUATION_ONLY_RE.match(remainder)


def should_force_break_after(text: str, match_end: int, delimiter: str) -> bool:
    if delimiter and text.startswith(delimiter, match_end):
        return False
    return should_insert_after_match(text, match_end)


def enforce_required_breaks(text: str, delimiter: str) -> str:
    def replacer(match: re.Match) -> str:
        if should_force_break_after(text, match.end(), delimiter):
            return match.group(0) + delimiter
        return match.group(0)

    return re.sub(r"を", replacer, text)


def insert_delimiters(text: Any, delimiter: str = DEFAULT_DELIMITER) -> Any:
    if not isinstance(text, str):
        return text

    def replacer(match: re.Match) -> str:
        matched_text = match.group(0)
        if delimiter and text.startswith(delimiter, match.end()):
            return matched_text
        if not should_insert_after_match(text, match.end()):
            return matched_text
        return matched_text + delimiter

    processed = BREAK_PATTERN.sub(replacer, text)
    processed = enforce_required_breaks(processed, delimiter)
    return postprocess_ellipses(processed, delimiter)


def with_joiners_between_non_breaks(text: str, break_marker: str, joiner: str) -> str:
    output = []
    index = 0

    while index < len(text):
        if text.startswith(break_marker, index):
            next_index = index + len(break_marker)
            previous_char = output[-1] if output else ""
            next_char = text[next_index] if next_index < len(text) else ""
            should_join_across_marker = (
                previous_char
                and next_char
                and (is_fullwidth_punctuation(previous_char) or is_fullwidth_punctuation(next_char))
                and can_join_char_pair(previous_char, next_char)
            )
            if should_join_across_marker:
                output.append(joiner)
            index += len(break_marker)
            continue

        output.append(text[index])
        next_index = index + 1
        should_join = (
            next_index < len(text)
            and can_join_char_pair(text[index], text[next_index])
            and not text.startswith(break_marker, next_index)
        )
        if should_join:
            output.append(joiner)
        index = next_index

    return "".join(output)


def is_fullwidth_punctuation(char: str) -> bool:
    return char in JAPANESE_JOINABLE_PUNCTUATION or (
        unicodedata.category(char).startswith("P") and unicodedata.east_asian_width(char) in {"F", "W"}
    )


def is_japanese_script_char(char: str) -> bool:
    return bool(JAPANESE_JOINABLE_CHAR_RE.fullmatch(char))


def is_japanese_joinable_char(char: str) -> bool:
    return is_japanese_script_char(char) or is_fullwidth_punctuation(char)


def can_join_char_pair(left: str, right: str) -> bool:
    return (
        not left.isspace()
        and left not in WHITESPACE_CHARS
        and not right.isspace()
        and right not in WHITESPACE_CHARS
        and not is_fullwidth_punctuation(left)
        and (
            (is_japanese_script_char(left) and is_japanese_script_char(right))
            or (is_japanese_script_char(left) and is_fullwidth_punctuation(right))
        )
    )


def insert_word_joiners(
    text: Any,
    joiner: str = WORD_JOINER,
    break_marker: str = WORD_JOINER_BREAK_MARKER,
) -> Any:
    if not isinstance(text, str):
        return text

    if break_marker in text:
        raise ValueError("The temporary word-joiner break marker already exists in the source text.")

    marked_text = insert_delimiters(text, break_marker)
    return with_joiners_between_non_breaks(marked_text, break_marker, joiner)


def target_columns(sheet) -> list[int]:
    columns = []
    for cell in sheet[1]:
        header = str(cell.value).strip().lower() if cell.value is not None else ""
        if header in TARGET_HEADERS:
            columns.append(cell.column)
    return columns


def process_workbook_values(workbook, transform) -> int:
    processed_count = 0

    for sheet in workbook.worksheets:
        for column in target_columns(sheet):
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=column)
                new_value = transform(cell.value)
                if new_value != cell.value:
                    cell.value = new_value
                    processed_count += 1

    return processed_count


def process_workbook(workbook, delimiter: str = DEFAULT_DELIMITER) -> int:
    return process_workbook_values(workbook, lambda value: insert_delimiters(value, delimiter))


def process_workbook_word_joiners(workbook, joiner: str = WORD_JOINER) -> int:
    return process_workbook_values(workbook, lambda value: insert_word_joiners(value, joiner))


def default_output_path(input_path: str | Path) -> Path:
    path = Path(input_path)
    return path.with_name(f"delimiters_added_{path.stem}{path.suffix}")


def default_word_joiner_output_path(input_path: str | Path) -> Path:
    path = Path(input_path)
    return path.with_name(f"word_joiners_added_{path.stem}{path.suffix}")


def convert_workbook(
    input_path: str | Path,
    output_path: str | Path | None = None,
    delimiter: str = DEFAULT_DELIMITER,
) -> tuple[Path, int]:
    input_path = Path(input_path)
    output_path = Path(output_path) if output_path is not None else default_output_path(input_path)

    workbook = load_workbook(input_path)
    processed_count = process_workbook(workbook, delimiter)
    workbook.save(output_path)
    return output_path, processed_count


def convert_workbook_word_joiners(
    input_path: str | Path,
    output_path: str | Path | None = None,
    joiner: str = WORD_JOINER,
) -> tuple[Path, int]:
    input_path = Path(input_path)
    output_path = Path(output_path) if output_path is not None else default_word_joiner_output_path(input_path)

    workbook = load_workbook(input_path)
    processed_count = process_workbook_word_joiners(workbook, joiner)
    workbook.save(output_path)
    return output_path, processed_count


def convert_uploaded_workbook(
    filename: str,
    raw_bytes: bytes,
    delimiter: str = DEFAULT_DELIMITER,
) -> tuple[Path, int]:
    workbook = load_workbook(io.BytesIO(raw_bytes))
    processed_count = process_workbook(workbook, delimiter)
    output_path = default_output_path(filename)
    workbook.save(output_path)
    return output_path, processed_count


def convert_uploaded_workbook_word_joiners(
    filename: str,
    raw_bytes: bytes,
    joiner: str = WORD_JOINER,
) -> tuple[Path, int]:
    workbook = load_workbook(io.BytesIO(raw_bytes))
    processed_count = process_workbook_word_joiners(workbook, joiner)
    output_path = default_word_joiner_output_path(filename)
    workbook.save(output_path)
    return output_path, processed_count


def breakpoint_positions(text: str) -> list[int]:
    return [match.end() for match in BREAK_PATTERN.finditer(text)]


def choose_balanced_breaks(text: str, line_count: int) -> list[int]:
    positions = breakpoint_positions(text)
    if not positions or line_count <= 1:
        return []

    target_len = len(text) / line_count
    breaks = []
    last_break = 0

    for index in range(1, line_count):
        target_position = target_len * index
        valid_breaks = [position for position in positions if position > last_break]
        if not valid_breaks:
            break

        best_break = min(valid_breaks, key=lambda position: abs(position - target_position))
        breaks.append(best_break)
        last_break = best_break

    return sorted(set(breaks))


def split_at_positions(text: str, positions: list[int]) -> list[str]:
    chunks = []
    previous = 0
    for position in positions:
        chunks.append(text[previous:position])
        previous = position
    chunks.append(text[previous:])
    return chunks


def polish_linebreaks(chunks: list[str]) -> list[str]:
    adjusted = chunks[:]

    for index in range(1, len(adjusted)):
        match = re.match(rf"^([{re.escape(LEADING_PUNCTUATION)}]{{1,3}})", adjusted[index])
        if not match:
            continue

        token = match.group(1)
        adjusted[index - 1] += token
        adjusted[index] = adjusted[index][len(token) :]

    return [chunk for chunk in adjusted if chunk]


def split_balanced_lines(text: str, line_count: int) -> list[str]:
    return polish_linebreaks(split_at_positions(text, choose_balanced_breaks(text, line_count)))


def prompt_line_count() -> int:
    try:
        line_count = int(input("How many lines would you like to split it into? ").strip())
        if line_count < 1:
            raise ValueError
    except ValueError:
        print("Invalid line count. Defaulting to 2.")
        return 2
    return line_count


def run_processing_mode() -> None:
    delimiter_input = input("Enter a delimiter (press Enter for invisible ZWSP '\\u200B'): ").strip()
    delimiter = delimiter_input if delimiter_input else DEFAULT_DELIMITER
    preview_symbol = "[ZWSP]" if delimiter == DEFAULT_DELIMITER else delimiter

    print(f"Using delimiter: {delimiter!r}")
    print(f"Preview: 日本語{preview_symbol}テキスト")

    files = get_colab_files()
    if files is not None:
        print("\nPlease upload your Excel file:")
        uploaded = files.upload()
        if not uploaded:
            raise RuntimeError("No file uploaded.")
        filename = next(iter(uploaded))
        output_path, processed_count = convert_uploaded_workbook(filename, uploaded[filename], delimiter)
        files.download(str(output_path))
    else:
        input_path = input("Enter local filename: ").strip()
        output_path, processed_count = convert_workbook(input_path, delimiter=delimiter)

    print(f"Done! Processed {processed_count} cells.")
    print(f"File saved as: {output_path}")


def run_splitting_mode() -> None:
    text = input("Paste the Japanese text segment:\n").strip()
    line_count = prompt_line_count()
    chunks = split_balanced_lines(text, line_count)

    if len(chunks) == 1 and chunks[0] == text:
        print("No suitable breakpoints found in the text.")
        print(f"Original: {text}")
        return

    print("\nSuggested linebreaks:\n")
    for index, chunk in enumerate(chunks, start=1):
        print(f"{index:02d}: {chunk}")


def run_word_joiner_mode() -> None:
    joiner_input = input("Enter a joiner (press Enter for WORD JOINER '\\u2060'): ").strip()
    joiner = joiner_input if joiner_input else WORD_JOINER
    preview_symbol = "[WJ]" if joiner == WORD_JOINER else joiner

    print(f"Using joiner: {joiner!r}")
    print(f"Preview: 日{preview_symbol}本{preview_symbol}語")

    files = get_colab_files()
    if files is not None:
        print("\nPlease upload your Excel file:")
        uploaded = files.upload()
        if not uploaded:
            raise RuntimeError("No file uploaded.")
        filename = next(iter(uploaded))
        output_path, processed_count = convert_uploaded_workbook_word_joiners(filename, uploaded[filename], joiner)
        files.download(str(output_path))
    else:
        input_path = input("Enter local filename: ").strip()
        output_path, processed_count = convert_workbook_word_joiners(input_path, joiner=joiner)

    print(f"Done! Processed {processed_count} cells.")
    print(f"File saved as: {output_path}")


def main() -> None:
    print(
        "Choose a mode:\n"
        "  1. Insert delimiters into an Excel file (Processing)\n"
        "  2. Linebreak a Japanese text segment into balanced chunks (Splitting)\n"
        "  3. Insert word joiners everywhere except regex breakpoints\n"
    )

    mode = input("Enter 1, 2, or 3 (default: 1): ").strip() or "1"
    if mode == "1":
        run_processing_mode()
    elif mode == "2":
        run_splitting_mode()
    elif mode == "3":
        run_word_joiner_mode()
    else:
        print("Invalid mode selected.")


if __name__ == "__main__":
    main()

